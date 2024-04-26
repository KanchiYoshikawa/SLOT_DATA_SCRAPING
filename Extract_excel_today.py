"""
画像から機械割を出すプログラム
このプログラムでは全てのホールのすべての日にちのデータに対して処理を行う。
"""

import time
import os
import datetime
import openpyxl
import cv2
import numpy as np
import glob
import logging

path = './Extract_excel_today_log.txt'

f = open(path, mode='w')


def adjust_len(ws1, input_column="C"):
    for col in ws1.columns:
        max_length = 0
        column = col[0].column

        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        adjusted_width = (max_length) * 2
        ws1.column_dimensions["A"].width = adjusted_width
        columns = str(input_column) + str(len(col))
        ws1.auto_filter.ref = "A1:" + columns
        break

root = "./data/"

Hole_names = glob.glob(root + "*")



for Hole_name in Hole_names:
    Hole_dir = Hole_name
    print(Hole_dir)
    f.write(Hole_dir)
    f.write("\n")
    
    #ホールフォルダにあるフォルダの名前を取得
    path = Hole_dir
    Hole_dir = os.listdir(path)
    
    """
    Day_names = [f for f in Hole_dir if os.path.isdir(os.path.join(path, f))]
    print(Day_names)
    """
    
    dt_now = datetime.datetime.today() - datetime.timedelta(days=1)
    today = datetime.datetime.strftime(dt_now, '%Y-%m-%d')
    Day_names = [today]
    
    for Day_name in Day_names:
        Day_dir = path + "/" + Day_name
        Day = Day_name
        print(Day)
        f.write(Day)
        f.write("\n")
        
        medal_sum = 0
        game_sum = 0
        
        Month = int(Day[-5:-3])
        #Excelファイルの定義
        excel_name = Hole_name + "/" + str(Month) + "月解析結果一覧.xlsx"

        if os.path.exists(excel_name) == True:
            #Excelを開いてsheetの定義
            wb = openpyxl.load_workbook(excel_name)
            sheet_medal = wb[wb.sheetnames[0]]
            sheet_game = wb[wb.sheetnames[1]]
            sheet_output = wb[wb.sheetnames[2]]

            #何行目に書き込むかを確認する
            rows_exsist = int(sheet_medal.max_column) + 1
            last_row = sheet_medal.cell(row=1,column=rows_exsist).coordinate

            #日付の取得
            output_day = int(Day[-2:])

            output_cell = last_row
            sheet_medal[output_cell] = str(int(output_day)) + ("日")
            sheet_game[output_cell] = str(int(output_day)) + ("日")
            sheet_output[output_cell] = str(int(output_day)) + ("日")

            #機種名の取得
            path = Hole_name + ("/") + Day
            try:
                Machine_names = os.listdir(path)
                Machine_names = [f for f in Machine_names if os.path.isdir(os.path.join(path, f))]
            except:
                print("フォルダがありません")
                f.write("フォルダがありません")
                f.write("\n")
                break


            #ファイルの名前（日付）から曜日を抽出
            dt_now = Day
            dt_now = datetime.datetime.strptime(dt_now,'%Y-%m-%d')
            week = dt_now.isoweekday()
            #print(week)

            #Excelの何行目に記入するかの定義
            Excel_i = 1


            #取得した機種名全てで実行
            for Machine_name in Machine_names:
                path = Hole_name + ("/") + Day + ("/") + Machine_name
                Graphs = os.listdir(path)
                Graphs = [f for f in Graphs if os.path.isfile(os.path.join(path, f))]

                #取得したグラフ（pngファイル）全てで実行
                for Graph in Graphs:
                    
                    #ゲーム数と差枚数の初期化
                    game_total = 0
                    medal_output = 0
                    
                    
                    path = Hole_name + ("/") + Day + ("/") + Machine_name + ("/") + str(Graph)
                    #print(path)
                    if Graph[-4:] == ".png":
                        #画像を読み込む。
                        img_array = np.fromfile(path, dtype=np.uint8)
                        img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
                        #img = cv2.imread(path)

                        if week == 1:
                            lower_color = np.array([0, 50, 255])
                            upper_color = np.array([0, 51, 255])
                        #火曜日
                        elif week == 2:
                            lower_color = np.array([255, 0, 153])
                            upper_color = np.array([255, 0, 154])
                        #水曜日
                        elif week == 3:
                            lower_color = np.array([204, 153, 0])
                            upper_color = np.array([204, 154, 0])
                        #木曜日
                        elif week == 4:
                            lower_color = np.array([204, 0, 255])
                            upper_color = np.array([205, 0, 255])
                        #金曜日
                        elif week == 5:
                            lower_color = np.array([255, 0, 0])
                            upper_color = np.array([255, 0, 1])
                        #土曜日
                        elif week == 6:
                            lower_color = np.array([0, 153, 204])
                            upper_color = np.array([0, 154, 204])
                        #日曜日
                        elif week == 7:
                            lower_color = np.array([0, 133, 0])
                            upper_color = np.array([0, 134, 0])


                        # 指定した色に基づいたマスク画像の生成
                        try:
                            mask = cv2.inRange(img, lower_color, upper_color)
                            indices = np.dstack(np.where(mask == 255))
                            in_shape = indices.shape
                            indices = indices.reshape(in_shape[1], 2)

                            #x,yの順になるように配列を並び替える
                            indices2 = indices[:, ::-1]

                            #x軸の値でソートする
                            indices3 = indices2[np.argsort(indices2[:, 0])]
                            output = cv2.bitwise_and(img, img, mask = mask)

                            #xとyの距離を抜き出す
                            #0点の位置(50, 118)
                            if len(indices3)>0:
                                game_total = indices3[-1, 0] - 50
                                medal_output = 118 - indices3[-1, 1]

                            #Excelへの記入
                            sheet_medal[output_cell[0] + str(Excel_i+1)] = int(medal_output * 57.4712644)

                            sheet_game[output_cell[0] + str(Excel_i+1)] = int(game_total * 87.488)

                            sheet_output[output_cell[0] + str(Excel_i+1)] = 100 + int(medal_output * 57.4712644)/int(game_total * 87.488)/3*100
                            
                            medal_sum += sheet_medal["C" + str(Excel_i+1)].value
                            game_sum += sheet_game["C" + str(Excel_i+1)].value
                            
                        except:
                            game_total = 0
                            medal_output = 0

                            #Excelへの記入
                            sheet_medal[output_cell[0] + str(Excel_i+1)] = 0
                            sheet_game[output_cell[0] + str(Excel_i+1)] = 0
                            sheet_output[output_cell[0] + str(Excel_i+1)] = 0

                        Excel_i += 1

            adjust_len(sheet_medal, output_cell[0])
            adjust_len(sheet_game, output_cell[0])
            adjust_len(sheet_output, output_cell[0])
            
            #Excelへの記入
            sheet_medal[output_cell[0] + str(Excel_i+1)] = medal_sum
            sheet_game[output_cell[0] + str(Excel_i+1)] = game_sum
            sheet_output[output_cell[0] + str(Excel_i+1)] = 100 + medal_sum / game_sum / 3 * 100

        else:
            wb = openpyxl.Workbook()

            #差枚数のシートの作成
            sheet_medal = wb.create_sheet('差枚数', 0)
            sheet_medal["A1"] = "機種名"
            sheet_medal["B1"] = "台番号"

            #総回転数のシートの作成
            sheet_game = wb.create_sheet("総回転数", 1)
            sheet_game["A1"] = "機種名"
            sheet_game["B1"] = "台番号"

            #機械割のシートの作成
            sheet_output = wb.create_sheet('機械割', 2)
            sheet_output["A1"] = "機種名"
            sheet_output["B1"] = "台番号"

            #機種名の取得
            path = Hole_name + ("/") + Day
            try:
                Machine_names = os.listdir(path)
                Machine_names = [f for f in Machine_names if os.path.isdir(os.path.join(path, f))]
            except:
                print("フォルダがありません")
                f.write("フォルダがありません")
                f.write("\n")
                break


            #日付の取得
            output_day = int(Day[-2:])
            sheet_medal["C1"] = str(int(output_day)) + ("日")
            sheet_game["C1"] = str(int(output_day)) + ("日")
            sheet_output["C1"] = str(int(output_day)) + ("日")

            #ファイルの名前（日付）から曜日を抽出
            dt_now = Day
            dt_now = datetime.datetime.strptime(dt_now,'%Y-%m-%d')
            week = dt_now.isoweekday()
            #print(week)

            #Excelの何行目に記入するかの定義
            Excel_i = 1


            #取得した機種名全てで実行
            for Machine_name in Machine_names:
                path = Hole_name + ("/") + Day + ("/") + Machine_name
                Graphs = os.listdir(path)
                Graphs = [f for f in Graphs if os.path.isfile(os.path.join(path, f))]

                #取得したグラフ（pngファイル）全てで実行
                for Graph in Graphs:
                    
                    #ゲーム数と差枚数の初期化
                    game_total = 0
                    medal_output = 0
                    
                    if Graph[-4:] == ".png":
                        path = Hole_name + ("/") + Day + ("/") + Machine_name + ("/") + str(Graph)
                        #画像を読み込む。
                        img_array = np.fromfile(path, dtype=np.uint8)
                        img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
                        #img = cv2.imread(path)

                        if week == 1:
                            lower_color = np.array([0, 50, 255])
                            upper_color = np.array([0, 51, 255])
                        #火曜日
                        elif week == 2:
                            lower_color = np.array([255, 0, 153])
                            upper_color = np.array([255, 0, 154])
                        #水曜日
                        elif week == 3:
                            lower_color = np.array([204, 153, 0])
                            upper_color = np.array([204, 154, 0])
                        #木曜日
                        elif week == 4:
                            lower_color = np.array([204, 0, 255])
                            upper_color = np.array([205, 0, 255])
                        #金曜日
                        elif week == 5:
                            lower_color = np.array([255, 0, 0])
                            upper_color = np.array([255, 0, 1])
                        #土曜日
                        elif week == 6:
                            lower_color = np.array([0, 153, 204])
                            upper_color = np.array([0, 154, 204])
                        #日曜日
                        elif week == 7:
                            lower_color = np.array([0, 133, 0])
                            upper_color = np.array([0, 134, 0])


                        # 指定した色に基づいたマスク画像の生成
                        try:
                            mask = cv2.inRange(img, lower_color, upper_color)
                            indices = np.dstack(np.where(mask == 255))
                            in_shape = indices.shape
                            indices = indices.reshape(in_shape[1], 2)

                            #x,yの順になるように配列を並び替える
                            indices2 = indices[:, ::-1]

                            #x軸の値でソートする
                            indices3 = indices2[np.argsort(indices2[:, 0])]
                            output = cv2.bitwise_and(img, img, mask = mask)

                            #xとyの距離を抜き出す
                            #0点の位置(50, 118)
                            if len(indices3)>0:
                                game_total = indices3[-1, 0] - 50
                                medal_output = 118 - indices3[-1, 1]

                            #Excelへの記入
                            sheet_medal["A" + str(Excel_i+1)] = Machine_name
                            sheet_medal["B" + str(Excel_i+1)] = int(Graph[:-4])
                            sheet_medal["C" + str(Excel_i+1)] = int(medal_output * 57.4712644)

                            sheet_game["A" + str(Excel_i+1)] = Machine_name
                            sheet_game["B" + str(Excel_i+1)] = int(Graph[:-4])
                            sheet_game["C" + str(Excel_i+1)] = int(game_total * 87.488)

                            sheet_output["A" + str(Excel_i+1)] = Machine_name
                            sheet_output["B" + str(Excel_i+1)] = int(Graph[:-4])
                            sheet_output["C" + str(Excel_i+1)] = 100 + int(medal_output * 57.4712644)/int(game_total * 87.488)/3*100
                            
                            medal_sum += sheet_medal["C" + str(Excel_i+1)].value
                            game_sum += sheet_game["C" + str(Excel_i+1)].value
                            
                        except:
                            game_total = 0
                            medal_output = 0

                            #Excelへの記入
                            sheet_medal["A" + str(Excel_i+1)] = Machine_name
                            sheet_medal["B" + str(Excel_i+1)] = int(Graph[:-4])
                            sheet_medal["C" + str(Excel_i+1)] = 0

                            sheet_game["A" + str(Excel_i+1)] = Machine_name
                            sheet_game["B" + str(Excel_i+1)] = int(Graph[:-4])
                            sheet_game["C" + str(Excel_i+1)] = 0

                            sheet_output["A" + str(Excel_i+1)] = Machine_name
                            sheet_output["B" + str(Excel_i+1)] = int(Graph[:-4])
                            sheet_output["C" + str(Excel_i+1)] = 0


                        Excel_i += 1
                        

            adjust_len(sheet_medal)
            adjust_len(sheet_game)
            adjust_len(sheet_output)
            
            #Excelへの記入
            sheet_medal["C" + str(Excel_i+1)] = medal_sum
            sheet_game["C" + str(Excel_i+1)] = game_sum
            sheet_output["C" + str(Excel_i+1)] = 100 + medal_sum / game_sum / 3 * 100


        wb.save(excel_name)

f.close()